# -*- coding: utf-8 -*-
"""P0 적재: articles.xlsx → 표준 기사 JSONL + 보조 라벨 사이드카 + 제외 기록.

표준 기사 코어 스키마(크롤링 현실 기준): article_id · title · posted_date · url · text
- article_id는 URL 해시("A" + sha1[:8]) — 기사 추가·재정렬·재실행에도 불변
- 분류 4(크롤링 오류)는 적재하지 않되 articles_excluded.jsonl에 기록 (전수 회계)
- 본문 text는 그대로 통과시킨다 — 정제는 P1 소관 (원본 보존 원칙)

사용:
    python -m src.p0_load --input D:/part1/articles.xlsx --outdir data
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

REQUIRED_COLUMNS = ["title", "posted date", "url", "text", "temporary classification"]
VALID_CLASSES = (1, 2, 3, 4)
EXCLUDED_CLASS = 4  # 크롤링 오류(본문 하단 잘림) — 파이프라인 미투입 확정

AUX_FIELDNAMES = ["url", "article_id", "news_source", "query", "journalist", "temp_class"]


def make_article_id(url: str) -> str:
    """URL 기반 결정적 기사 ID. 앞뒤 공백·끝 슬래시 차이는 같은 기사로 본다."""
    normalized = url.strip().rstrip("/")
    return "A" + hashlib.sha1(normalized.encode("utf-8")).hexdigest()[:8]


def normalize_posted_date(value) -> str:
    """xlsx의 datetime/date/문자열 작성일을 'YYYY-MM-DD' 문자열로 통일."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = "" if value is None else str(value).strip()
    m = re.match(r"^(\d{4})[-./](\d{1,2})[-./](\d{1,2})", s)
    if m:
        y, mo, d = (int(g) for g in m.groups())
        return date(y, mo, d).isoformat()
    raise ValueError(f"posted date 해석 불가: {value!r}")


def read_rows(xlsx_path: Path) -> list[dict]:
    """첫 시트를 헤더 기준 dict 행 목록으로 읽는다. 완전 빈 행은 무시."""
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        missing = [c for c in REQUIRED_COLUMNS if c not in header]
        if missing:
            raise ValueError(f"필수 컬럼 누락: {missing} / 실제 헤더: {header}")
        idx = {h: i for i, h in enumerate(header) if h}
        rows = []
        for r in rows_iter:
            if r is None or all(v is None or str(v).strip() == "" for v in r):
                continue
            rows.append({h: (r[i] if i < len(r) else None) for h, i in idx.items()})
        return rows
    finally:
        wb.close()


def transform(rows: list[dict]) -> tuple[list[dict], list[dict], list[dict]]:
    """행 검증 + 표준화. 반환: (articles, aux_labels, excluded).

    - 검증 실패가 하나라도 있으면 전체 실패(ValueError에 전 행 사유 나열) — 부분 산출 금지
    - 인바리언트: len(articles) + len(excluded) == len(rows)
    """
    errors: list[str] = []
    articles: list[dict] = []
    aux: list[dict] = []
    excluded: list[dict] = []
    seen_urls: dict[str, int] = {}
    seen_ids: dict[str, int] = {}

    for lineno, row in enumerate(rows, start=2):  # xlsx 기준 행 번호(1행은 헤더)
        row_errors: list[str] = []

        url = str(row.get("url") or "").strip()
        title = str(row.get("title") or "").strip()
        text_raw = row.get("text")
        text = "" if text_raw is None else str(text_raw)  # 원본 보존 — strip도 하지 않음

        if not url.lower().startswith("http"):
            row_errors.append(f"{lineno}행: url 비정상 {url!r}")
        if not title:
            row_errors.append(f"{lineno}행: title 비어 있음")
        if not text.strip():
            row_errors.append(f"{lineno}행: text 비어 있음")

        posted = None
        try:
            posted = normalize_posted_date(row.get("posted date"))
        except ValueError as e:
            row_errors.append(f"{lineno}행: {e}")

        temp_class = None
        tc_raw = row.get("temporary classification")
        try:
            temp_class = int(str(tc_raw).strip())
            if temp_class not in VALID_CLASSES:
                raise ValueError
        except (TypeError, ValueError):
            row_errors.append(f"{lineno}행: temporary classification 비정상 {tc_raw!r}")

        if url:
            if url in seen_urls:
                row_errors.append(f"{lineno}행: url 중복 (최초 {seen_urls[url]}행)")
            else:
                seen_urls[url] = lineno

        if row_errors:
            errors.extend(row_errors)
            continue

        aid = make_article_id(url)
        provided_id = str(row.get("article ID") or "").strip()
        if provided_id and provided_id != aid:
            # xlsx에 기재된 ID가 있으면 URL 계산값과 일치해야 한다 (수기 오류·복사 실수 방어)
            errors.append(f"{lineno}행: article ID 불일치 — 기재 {provided_id!r} ≠ URL 계산 {aid!r}")
            continue
        if aid in seen_ids:
            errors.append(f"{lineno}행: article_id 충돌 {aid} (최초 {seen_ids[aid]}행)")
            continue
        seen_ids[aid] = lineno

        aux.append({
            "url": url,
            "article_id": aid,
            "news_source": "" if row.get("news source") is None else str(row["news source"]).strip(),
            "query": "" if row.get("query") is None else str(row["query"]).strip(),
            "journalist": "" if row.get("journalist") is None else str(row["journalist"]).strip(),
            "temp_class": temp_class,
        })

        if temp_class == EXCLUDED_CLASS:
            excluded.append({
                "article_id": aid,
                "url": url,
                "title": title,
                "temp_class": temp_class,
                "reason_code": "CRAWL_ERROR_TRUNCATED",
                "reason": "크롤링 오류(본문 하단 잘림) — 원문 대조로 확인, 파이프라인 미투입",
            })
        else:
            articles.append({
                "article_id": aid,
                "title": title,
                "posted_date": posted,
                "url": url,
                "text": text,
            })

    if errors:
        raise ValueError("P0 적재 검증 실패:\n" + "\n".join(errors))

    assert len(articles) + len(excluded) == len(rows), "전수 회계 인바리언트 위반"
    return articles, aux, excluded


def write_outputs(outdir: Path, articles: list[dict], aux: list[dict], excluded: list[dict]) -> dict[str, Path]:
    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    paths = {
        "articles": outdir / "articles.jsonl",
        "aux": outdir / "aux_labels.csv",
        "excluded": outdir / "articles_excluded.jsonl",
    }
    with open(paths["articles"], "w", encoding="utf-8", newline="\n") as f:
        for a in articles:
            f.write(json.dumps(a, ensure_ascii=False) + "\n")
    with open(paths["excluded"], "w", encoding="utf-8", newline="\n") as f:
        for e in excluded:
            f.write(json.dumps(e, ensure_ascii=False) + "\n")
    # 사이드카는 Excel에서 바로 열어보는 용도가 커서 BOM 포함(utf-8-sig)으로 쓴다
    with open(paths["aux"], "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=AUX_FIELDNAMES)
        w.writeheader()
        w.writerows(aux)
    return paths


def main(argv=None) -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    p = argparse.ArgumentParser(description="P0 적재: articles.xlsx → 표준 기사 JSONL + 사이드카 + 제외 기록")
    p.add_argument("--input", type=Path, default=Path("D:/part1/articles.xlsx"), help="입력 xlsx 경로")
    p.add_argument("--outdir", type=Path, default=Path("data"), help="산출물 디렉터리")
    args = p.parse_args(argv)

    rows = read_rows(args.input)
    articles, aux, excluded = transform(rows)
    paths = write_outputs(args.outdir, articles, aux, excluded)

    dist: dict[int, int] = {}
    for r in aux:
        dist[r["temp_class"]] = dist.get(r["temp_class"], 0) + 1
    print(f"입력 {len(rows)}건 = 적재 {len(articles)}건 + 제외 {len(excluded)}건 — 전수 회계 OK")
    print(f"분류 분포: {dict(sorted(dist.items()))}")
    for name, path in paths.items():
        print(f"  {name}: {path}")


if __name__ == "__main__":
    main()
